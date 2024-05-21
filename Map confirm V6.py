import os
import re
import sys
from PyQt5.QtWidgets import QApplication, QWidget, QLabel, QVBoxLayout, QPushButton, QFileDialog, QLineEdit, QColorDialog, QMessageBox, QCheckBox
from PyQt5.QtGui import QPixmap, QFont, QImage, QPainter, QColor, QIcon
from PyQt5.QtCore import Qt, QRect
import qtmodern.styles
import qtmodern.windows
import openpyxl
from openpyxl.styles import PatternFill

# 判斷是否為打包後的執行檔
if getattr(sys, 'frozen', False):
    application_path = sys._MEIPASS
else:
    application_path = os.path.dirname(os.path.abspath(__file__))

# 設定圖示路徑
icon_path = os.path.join(application_path, 'format.ico')

class ImageClassifier(QWidget):
    def __init__(self):
        super().__init__()

        # 設定視窗標題和圖示
        self.setWindowTitle('Map Confirm')
        app.setWindowIcon(QIcon(icon_path))
        self.resize(800, 600)
        
        # 建立導入MAP圖按鈕
        self.importBtn = QPushButton('導入MAP圖', self)
        self.importBtn.clicked.connect(self.importMap)
        
        # 建立選擇確認照片資料夾按鈕
        self.confirmPathBtn = QPushButton('選擇確認照片的資料夾', self)
        self.confirmPathBtn.clicked.connect(self.selectConfirmPath)
        
        # 建立執行按鈕
        self.runSaveBtn = QPushButton('執行', self)
        self.runSaveBtn.clicked.connect(self.executeSave)

        # 建立另存新檔按鈕
        self.runSaveAsBtn = QPushButton('另存新檔', self)
        self.runSaveAsBtn.clicked.connect(self.executeSaveAs)
        
        # 建立設定填充顏色按鈕
        self.colorBtn = QPushButton('設定填充顏色', self)
        self.colorBtn.clicked.connect(self.setFillColor)

        # 建立不要覆蓋已存在的Issue code的複選框
        self.overwriteCheckBox = QCheckBox('不要覆蓋已存在的Issue code', self)
        self.overwriteCheckBox.setChecked(True)
        
        # 建立MAP檔案路徑標籤
        self.pathLabel = QLineEdit(self)
        self.pathLabel.hide()
        
        # 建立確認照片資料夾路徑標籤
        self.confirmPathLabel = QLineEdit(self)
        self.confirmPathLabel.hide()
        
        # 建立圖片標籤
        self.imageLabel = QLabel(self)
        
        # 設定預設的填充顏色
        self.start_color = "bc8f8f"
        self.end_color = "bc8f8f"
        
        # 建立佈局並將元件加入佈局
        layout = QVBoxLayout()
        layout.addWidget(self.importBtn)
        layout.addWidget(self.pathLabel)
        layout.addWidget(self.confirmPathBtn)
        layout.addWidget(self.confirmPathLabel)
        layout.addWidget(self.colorBtn)
        layout.addWidget(self.overwriteCheckBox)
        layout.addWidget(self.runSaveAsBtn)
        layout.addWidget(self.runSaveBtn)
        layout.addWidget(self.imageLabel)
        self.setLayout(layout)
        
        # 檢查是否顯示執行按鈕
        self.checkShowRunButton()
        
    def importMap(self):
        # 開啟檔案對話框選擇MAP檔案
        path, _ = QFileDialog.getOpenFileName(self, '選擇MAP檔案', filter="Excel files (*.xlsx)")
        if path:
            self.pathLabel.setText(path)
            self.pathLabel.show()
            self.displayMap(path)
            self.checkShowRunButton()
    
    def selectConfirmPath(self):
        # 開啟資料夾對話框選擇確認照片的資料夾
        folder_path = QFileDialog.getExistingDirectory(self, '選擇確認照片的資料夾')
        if folder_path:
            self.confirmPathLabel.setText(folder_path)
            self.confirmPathLabel.show()
            self.checkShowRunButton()

    def checkShowRunButton(self):
        # 檢查是否顯示執行按鈕
        if self.pathLabel.text() and self.confirmPathLabel.text():
            self.runSaveBtn.show()
            self.runSaveAsBtn.show()
        else:
            self.runSaveBtn.hide()
            self.runSaveAsBtn.hide()

    def setFillColor(self):
        # 設定填充顏色
        color = QColorDialog.getColor()
        if color.isValid():
            rgb_color = f'{color.red():02X}{color.green():02X}{color.blue():02X}'
            self.start_color = rgb_color
            self.end_color = rgb_color

    def executeSave(self):
        # 執行儲存
        self.executeCore(self.pathLabel.text())

    def executeSaveAs(self):
        # 另存新檔
        save_path, _ = QFileDialog.getSaveFileName(self, '另存新檔', filter="Excel files (*.xlsx)")
        if save_path:
            # 載入目前的工作簿並另存為新檔案，不做任何修改
            workbook = openpyxl.load_workbook(self.pathLabel.text())
            workbook.save(save_path)
            workbook.close()
            
            # 更新主檔案路徑標籤以反映新的檔案位置，不改變確認路徑
            self.pathLabel.setText(save_path)
            self.displayMap(save_path)  # 可選：顯示新檔案路徑的地圖

    def executeCore(self, save_path):
        # 執行核心功能
        confirm_folder_path = self.confirmPathLabel.text()
        if os.path.exists(confirm_folder_path):
            xy_points = self.extract_xy_points(confirm_folder_path)
            adjusted_xy_points = [(str(int(x) + 4), str(int(y) + 2)) for x, y in xy_points]

            workbook = openpyxl.load_workbook(self.pathLabel.text())
            worksheet = workbook.active

            for x_point, y_point in adjusted_xy_points:
                try:
                    x_point = int(x_point)
                    y_point = int(y_point)
                    if x_point >= 1 and y_point >= 1:
                        cell = worksheet.cell(row=y_point, column=x_point)
                        if not self.overwriteCheckBox.isChecked() or cell.value is None:
                            cell.fill = PatternFill(start_color=self.start_color, end_color=self.end_color, fill_type="solid")
                except ValueError:
                    pass

            workbook.save(save_path)
            workbook.close()

            self.displayMap(save_path)

    def extract_xy_points(self, folder_path):
        # 從照片檔名提取XY座標
        xy_points = []
        for filename in os.listdir(folder_path):
            if filename.lower().endswith(('.jpg', '.jpeg', '.png')):
                # 現有的模式
                if "-" in filename:
                    match = re.search(r'-\d+_(\d+)_(\d+)_', filename)
                else:
                    match = re.search(r'[^_]*_[^_]*_([0-9]+)_([0-9]+)_', filename)

                # 新的模式，用於像是 "KK4QR07_2_8_Un-reviewed_1.jpg" 這樣的檔名
                if not match:
                    match = re.search(r'[^_]*_([0-9]+)_([0-9]+)_', filename)

                if match:
                    try:
                        x_point = int(match.group(1))
                        y_point = int(match.group(2))
                    except ValueError:
                        raise ValueError(f"檔名中的 x_point 或 y_point 無效: {filename}")

                    xy_points.append((x_point, y_point))
                else:
                    raise ValueError(f"檔名不符合模式: {filename}")

        return xy_points

    def get_fill_color(self, cell):
        # 獲取儲存格的填充顏色
        fill = cell.fill
        if fill.fill_type == "solid" and hasattr(fill.start_color, 'rgb'):
            try:
                return QColor.fromRgb(*[int(fill.start_color.rgb[i:i+2], 16) for i in (2, 4, 6)])
            except ValueError:
                # 處理轉換錯誤
                return None
        else:
            return None

    def displayMap(self, excel_path):
        # 顯示地圖
        try:
            workbook = openpyxl.load_workbook(excel_path, data_only=True)
            worksheet = workbook.active
        except Exception as e:
            QMessageBox.critical(self, '錯誤', f'發生錯誤: {e}')

        # 找到有值的範圍
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        start_col = 3  # 從C欄開始

        # 找到最後一欄有值的位置
        end_col = max_col
        while end_col >= start_col and all(worksheet.cell(row, end_col).value is None for row in range(1, max_row + 1)):
            end_col -= 1

        # 在找到的最後一欄有值的位置後再往右增加一欄
        end_col += 1

        # 儲存顏色計數的字典
        color_counts = {}

        if end_col >= start_col:
            # 設定單元格大小
            cell_width = 8
            cell_height = 6
            
            # 計算圖片大小
            image_width = (end_col - start_col) * cell_width
            image_height = max_row * cell_height
            
            # 創建一張新的圖片
            image = QImage(image_width, image_height, QImage.Format_RGB32)
            image.fill(Qt.white)
            painter = QPainter(image)
            
            # 繪製表格，包括儲存格的底色
            for row in range(max_row):
                for col in range(start_col, end_col):
                    cell = worksheet.cell(row + 1, col)
                    cell_value = cell.value
                    cell_rect = QRect((col - start_col) * cell_width, row * cell_height, cell_width, cell_height)
                    
                    # 獲取儲存格的填充顏色
                    fill_color = self.get_fill_color(cell)
                    if fill_color:
                        painter.fillRect(cell_rect, fill_color)  # 使用填充顏色填充矩形
                        # 計算顏色出現的次數
                        color_hex = fill_color.name()[1:]  # 將QColor轉換為不帶'#'的十六進制字串
                        color_counts[color_hex] = color_counts.get(color_hex, 0) + 1
                    
                    painter.drawRect(cell_rect)
                    if cell_value:
                        painter.drawText(cell_rect, Qt.AlignCenter, str(cell_value))
            
            painter.end()
            
            # 顯示圖片
            pixmap = QPixmap.fromImage(image)
            self.imageLabel.setPixmap(pixmap)
            self.imageLabel.setScaledContents(True)  # 保持圖片的原始尺寸
            self.imageLabel.setAlignment(Qt.AlignCenter)
        else:
            self.imageLabel.clear()
            self.imageLabel.setText('沒有資料')

        # 清除現有數據和格式從A12開始
        for row in range(12, worksheet.max_row + 1):
            worksheet[f'A{row}'].value = None
            worksheet[f'A{row}'].fill = openpyxl.styles.PatternFill(fill_type=None)
            worksheet[f'B{row}'].value = None
            worksheet[f'B{row}'].fill = openpyxl.styles.PatternFill(fill_type=None)

        # 寫入顏色計數到Excel
        row_counter = 12  # 從第12行開始寫入
        for color, count in color_counts.items():
            color_cell = f'A{row_counter}'
            count_cell = f'B{row_counter}'

            worksheet[color_cell].fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type="solid")
            worksheet[count_cell].value = count

            row_counter += 1

        # 創建 Arial 字體，大小為 9 的 Font 對象
        arial_font = openpyxl.styles.Font(name='微軟正黑體', size=9)

        # 設置 A 欄和 B 欄的單元格字體
        for row in range(12, row_counter):
            worksheet[f'A{row}'].font = arial_font
            worksheet[f'B{row}'].font = arial_font

        # 保存修改後的工作簿
        workbook.save(excel_path)
        workbook.close()

if __name__ == '__main__':
    app = QApplication(sys.argv)
    font = QFont("微軟正黑體", 9)
    app.setFont(font)
    qtmodern.styles.dark(app)

    window = ImageClassifier()
    win = qtmodern.windows.ModernWindow(window)
    win.show()

    # 獲取螢幕的寬度和高度
    screen = app.primaryScreen()
    screen_width = screen.geometry().width()
    screen_height = screen.geometry().height()

    # 計算視窗的左上角座標，使其在螢幕中央
    window_width = win.frameGeometry().width()
    window_height = win.frameGeometry().height()
    x = (screen_width - window_width) // 2
    y = (screen_height - window_height) // 2

    # 設定視窗位置
    win.move(x, y)

    sys.exit(app.exec_())
