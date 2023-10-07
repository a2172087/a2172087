import sys
sys.stdout.reconfigure(encoding='utf-8')
import os
import docx
import openpyxl

# Desktop Path
desktop_path = os.path.expanduser("~/Desktop")

# Read .docx file
file_path = os.path.join(desktop_path, '身分資料文件.docx')

# Create an empty dictionary to store names and IDs.
data = {'Name': [], 'ID': [], 'Last_name': [], 'First_name': []}

try:
    # Open.docx file
    doc = docx.Document(file_path)
    
    for paragraph in doc.paragraphs:
        line = paragraph.text.strip()
        # pass empty line
        if not line.strip():
            continue
        
        # ',' Ends
        if not line.endswith(','):
            raise ValueError('錯誤：身分資料文件中的某行不以逗號結尾。')
        
        parts = line.split(',')
        data['Name'].append(parts[0].strip())
        data['ID'].append(parts[1].strip())

    for name in data['Name']:
        name_parts = name.split()
        if len(name_parts) == 2:
            last_name, first_name = name_parts
        elif len(name_parts) == 1:
            last_name, first_name = name_parts[0], ''
        else:
            last_name, first_name = '', ''
        data['Last_name'].append(last_name)
        data['First_name'].append(first_name)

    data['ID'] = [id_value for id_value in data['ID'] if id_value.strip()]

    county_code_mapping = {
        'A': '臺北市',
        'B': '臺中市',
        'C': '基隆市',
        'D': '臺南市',
        'E': '高雄市',
        'F': '新北市',
        'G': '宜蘭縣',
        'H': '桃園市',
        'J': '新竹縣',
        'K': '苗栗縣',
        'L': '臺中縣',
        'M': '南投縣',
        'N': '彰化縣',
        'P': '雲林縣',
        'Q': '嘉義縣',
        'R': '臺南縣',
        'S': '高雄縣',
        'T': '屏東縣',
        'U': '花蓮縣',
        'V': '臺東縣',
        'X': '澎湖縣',
        'Y': '陽明山',
        'W': '金門縣',
        'Z': '連江縣',
        'I': '嘉義市',
        'O': '新竹市'
    }

    data['Born'] = []

    for id_value in data['ID']:
        if id_value[0] in county_code_mapping:
            birthplace = county_code_mapping[id_value[0]]
            data['Born'].append(birthplace)
        else:
            data['Born'].append('未知')

    sex_mapping = {
        '1': 'male',
        '2': 'Female'
    }

    data['Sex'] = []

    for id_value in data['ID']:
        if len(id_value) >= 2 and id_value[1] in sex_mapping:
            sex = sex_mapping[id_value[1]]
            data['Sex'].append(sex)
        else:
            data['Sex'].append('未知')

    # Create a new Excel workbook
    workbook = openpyxl.Workbook()

    # Select the default sheet
    sheet = workbook.active

    # Add headers
    sheet['A1'] = '姓'
    sheet['B1'] = '名'
    sheet['C1'] = '性別'
    sheet['D1'] = '身分證'
    sheet['E1'] = '戶籍地'

    # Populate data
    for i in range(len(data['Name'])):
        sheet.cell(row=i + 2, column=1, value=data['First_name'][i])
        sheet.cell(row=i + 2, column=2, value=data['Last_name'][i])
        sheet.cell(row=i + 2, column=3, value=data['Sex'][i])
        sheet.cell(row=i + 2, column=4, value=data['ID'][i])
        sheet.cell(row=i + 2, column=5, value=data['Born'][i])

    # Save the workbook to the desktop as Result.xlsx
    result_file_path = os.path.join(desktop_path, 'Result.xlsx')
    workbook.save(result_file_path)

    print(f'結果已成功保存到桌面上的Result.xlsx檔案中：{result_file_path}')

except FileNotFoundError:
    print(f'錯誤：找不到檔案 {file_path}')
except Exception as e:
    print(f'錯誤：{str(e)}')
