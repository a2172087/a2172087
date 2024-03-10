import sys
import os
import re
import openpyxl
from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, QFileDialog, QLineEdit, QMessageBox
from PyQt5.QtGui import QFont, QIcon
from PyQt5 import QtGui
import qtmodern.styles
import qtmodern.windows

app = QApplication(sys.argv)

# 取得桌面路徑
desktop_path = os.path.expanduser("~/Desktop")

# 檔案路徑
file_name = "standard_deviation_for_confidence.txt"
file_path = os.path.join(desktop_path, file_name)

# 讀取文件
with open(file_path, 'r') as file:
    lines = file.readlines()

# 移除 " confidence:" 字元以及其他處理
lines = [line.replace(" confidence:", "").replace(", #", "").replace(" ", ",").replace(",,", ",") for line in lines]

# 創建空的結果列表
result = []

# 定義C欄位的映射字典
c_value_mapping = {
    "!16": "Particle",
    "!15": "Over_kill",
    "!10": "Probe_Mark_Shift",
    "!09": "Foreign_material",
    "!07": "Process_Defect",
    "!2D": "Ugly_Die",
    "!25": "Bump_foreign_Material",
    "!28": "Bump_house_defect",
    "!0F": "Al_particle_out_of_PAD",
    "!1B": "Pad_discoloration",
    "!08": "Wafer_Scratch",
    "!26": "Missing_bump",
    "!18": "Al_particle"
}

# 定義B欄位的映射字典
b_value_mapping = {
    "@16": "Particle",
    "@15": "Over_kill",
    "@10": "Probe_Mark_Shift",
    "@09": "Foreign_material",
    "@07": "Process_Defect",
    "@2D": "Ugly_Die",
    "@25": "Bump_foreign_Material",
    "@28": "Bump_house_defect",
    "@0F": "Al_particle_out_of_PAD",
    "@1B": "Pad_discoloration",
    "@08": "Wafer_Scratch",
    "@26": "Missing_bump",
    "@18": "Al_particle"
}

# 解析每一行的數據
for line in lines:
    parts = line.split(',')
    if len(parts) >= 3:  # 檢查至少有3個元素 (Device, ADC, Check)
        device = parts[0]
        check = parts[1]
        adc = parts[2]

        # 檢查 C 欄位並替換值
        check = c_value_mapping.get(check.strip(), check)

        defect_values = parts[3:]

        # 創建一個字典來存儲 Defect_code
        defect_code_dict = {}
        for i, value in enumerate(defect_values):
            defect_code_dict[f"Defect_code_{i + 1}"] = value

        # 將所有資料整合到一個字典中
        data = {
            'Device': device,
            'After': adc,
            'Before': check,
            **defect_code_dict  # 將 Defect_code 欄位加入字典
        }

        result.append(data)

# 創建 Excel 文件
wb = openpyxl.Workbook()
ws = wb.active

# 添加表頭
header = list(result[0].keys())
ws.append(header)

# 添加數據行
for item in result:
    row = [item[key] for key in header]
    ws.append(row)

# 輸出 Excel 檔案
output_file_path = os.path.join(desktop_path, "standard_deviation_for_confidence.xlsx")
wb.save(output_file_path)

# 取得桌面路徑
desktop_path = os.path.expanduser("~/Desktop")
# 指定檔案名稱
file_name = "standard_deviation_for_confidence.xlsx"
# 組合完整的檔案路徑
file_path = os.path.join(desktop_path, file_name)
# 開啟Excel檔案
workbook = openpyxl.load_workbook(file_path)
# 選取要設定欄寬的工作表（Sheet）
workSheet = workbook['Sheet']

# 檢查 B 欄位並替換值
for row in workSheet.iter_rows(min_row=2, min_col=2, max_col=2):
    for cell in row:
        cell_value = cell.value
        if cell_value in b_value_mapping:
            cell.value = b_value_mapping[cell_value]

# 輸出 Excel 檔案
output_file_path = os.path.join(desktop_path, "standard_deviation_for_confidence.xlsx")
workbook.save(output_file_path)

class DataProcessingApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.init_ui()
        self.setWindowTitle('SDC')
        app.setWindowIcon(QtGui.QIcon('icon_tool.ico'))
        
    def init_ui(self):
        self.setWindowTitle("SDC")
        self.setGeometry(370, 350, 300, 300)

        # 新增按鈕 "Input config模型"
        self.input_defect_button = QPushButton("選擇 config.json 檔案", self)
        self.input_defect_button.setGeometry(20, 20, 250, 40)
        self.input_defect_button.clicked.connect(self.open_file_dialog)

        # 顯示選擇的檔案路徑的輸入欄
        self.file_path_label = QLineEdit(self)
        self.file_path_label.setGeometry(20, 80, 250, 30)
        self.file_path_label.setReadOnly(True)

        # 新增確認按鈕
        self.confirm_button = QPushButton("確認", self)
        self.confirm_button.setGeometry(20, 130, 250, 30)
        self.confirm_button.clicked.connect(self.process_file)

        self.file_path = ""
        self.number_1_values = []  # 存儲Number_1的值列表

    def open_file_dialog(self):
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly  # 讓選擇的檔案以只讀模式開啟
        file_path, _ = QFileDialog.getOpenFileName(self, "選擇檔案", "", "All Files (*)", options=options)

        if file_path:
            self.file_path = file_path  # 儲存選擇的檔案路徑
            self.file_path_label.setText(file_path)  # 顯示檔案路徑在輸入欄中

    def process_file(self):
        try:
            with open(self.file_path, 'r', encoding='utf-8') as file:
                lines = file.readlines()
                # 提取"description"的值
                descriptions = [line.strip().split('"description": ')[1] for line in lines if '"description": ' in line]
                Number_1 = '\n'.join(descriptions)  # 將所有描述以換行符號連接
                self.number_1_values = Number_1.split('\n')  # 將Number_1的值分割成列表
                # 顯示成功導入的視窗
                QMessageBox.information(self, "Successful", "數據成功導入")
                self.update_excel()  # 呼叫更新Excel的方法
                sys.exit()  # 關閉程式
        except Exception as e:
            print(f"錯誤: {str(e)}")

    def update_excel(self):
        if not self.number_1_values:
            return

        try:
            # 獲取用戶的桌面路徑
            desktop_path = os.path.expanduser("~/Desktop")

            # 拼接完整的Excel文件路徑
            excel_file_path = os.path.join(desktop_path, "standard_deviation_for_confidence.xlsx")

            # 打開Excel檔案
            workbook = openpyxl.load_workbook(excel_file_path)
            sheet = workbook.active

            # 獲取Number_1值的數量
            num_values = len(self.number_1_values)

            # 從D1開始，依次替換數值
            for i in range(num_values):
                cell = sheet.cell(row=1, column=4 + i)  # 從第1行，第4列開始
                cell.value = self.number_1_values[i]

            # 處理D1到Z1的儲存格
            for col in range(4, 27):  # D 到 Z 對應的列索引範圍
                cell = sheet.cell(row=1, column=col)
                cell_value = cell.value
                if cell_value:
                    # 使用正規表達式只保留英文字母並轉為小寫
                    new_value = re.sub(r'[^a-zA-Z]', '', cell_value).lower()

                    # 進行特定值的替換
                    if new_value == 'duglydie':
                        new_value = 'uglydie'
                    elif new_value == 'bumppmdiameteroutofspec':
                        new_value = 'ebumppmdiameteroutofspec'
                    elif new_value == 'falparticleoutofpad':
                        new_value = 'alparticleoutofpad'
                    elif new_value == 'bpaddiscoloration':
                        new_value = 'paddiscoloration'
                    elif new_value == 'csurfaceincomingdefect':
                        new_value = 'surfaceincomingdefect'
                    elif new_value == 'dsmallbump':
                        new_value = 'smallbump'
                    elif new_value == 'especialpmshfit':
                        new_value = 'specialpmshfit'
                    elif new_value == 'ediecrack':
                        new_value = 'diecrack'
                    elif new_value == 'fwafersurfacescratch':
                        new_value = 'wafersurfacescratch'
                    elif new_value == 'bflydie':
                        new_value = 'flydie'
                    elif new_value == 'baother':
                        new_value = 'aother'
                    elif new_value == 'bboverkill':
                        new_value = 'boverkill'
                    elif new_value == 'bcincompletedie':
                        new_value = 'cincompletedie'
                    elif new_value == 'airregularbump':
                        new_value = 'irregularbump'
                    elif new_value == 'ebumppmdiameteroutofspec':
                        new_value = 'bumppmdiameteroutofspec'
                    elif new_value == 'falparticleoutofpad':
                        new_value = 'alparticleoutofpad'
                    elif new_value == 'bpaddiscoloration':
                        new_value = 'paddiscoloration'
                    elif new_value == 'csurfaceincomingdefect':
                        new_value = 'surfaceincomingdefect'
                    elif new_value == 'dsmallbump':
                        new_value = 'smallbump'
                    elif new_value == 'especialpmshfit':
                        new_value = 'specialpmshfit'

                    # 將新值設定回儲存格
                    cell.value = new_value

            # 將B欄的英文全部轉為小寫，然後"_"刪除
            for row in range(2, sheet.max_row + 1):  # 從第二行開始處理
                cell = sheet.cell(row=row, column=2)  # B欄
                cell_value = cell.value
                if cell_value:
                    # 使用正規表達式只保留英文字母並轉為小寫
                    new_value = re.sub(r'[^a-zA-Z]', '', cell_value).lower()

                    # 將新值設定回儲存格
                    cell.value = new_value

            # 將C欄的英文全部轉為小寫，然後"_"刪除
            for row in range(2, sheet.max_row + 1):  # 從第二行開始處理
                cell = sheet.cell(row=row, column=3)  # C欄
                cell_value = cell.value
                if cell_value:
                    # 使用正規表達式只保留英文字母並轉為小寫
                    new_value = re.sub(r'[^a-zA-Z]', '', cell_value).lower()

                    # 將新值設定回儲存格
                    cell.value = new_value

            # 獲取D1到V1的標題
            header_row = sheet[1]

            # 創建一個字典，將標題列的值作為鍵，其對應的欄索引作為值
            header_dict = {cell.value: cell.column for cell in header_row}

            # 開始處理數據，從第二列開始
            row_num = 2
            while True:
                # 獲取C欄的值
                c_value = sheet.cell(row=row_num, column=3).value
                
                # 如果C欄的值為空，則停止處理
                if not c_value:
                    break
                
                # 尋找匹配的欄位
                matching_column = header_dict.get(c_value)
                
                if matching_column:
                    # 保留匹配欄位的值，刪除其他欄位的值
                    for col in range(4, 27):
                        if col != matching_column:
                            sheet.cell(row=row_num, column=col).value = None
                
                # 移動到下一列
                row_num += 1

            # 儲存Excel檔案，取代原始檔案
            desktop_path = os.path.expanduser("~/Desktop")
            output_file_path = os.path.join(desktop_path, "standard_deviation_for_confidence.xlsx")
            workbook.save(output_file_path)

            # 顯示成功導入的視窗
            #QMessageBox.information(self, "Successful", "資料編輯完成")
        except Exception as e:
            print(f"錯誤: {str(e)}")


def main():
    app = QApplication(sys.argv)
    font = QFont("微軟正黑體", 9)
    app.setFont(font)
    qtmodern.styles.dark(app)

    window = DataProcessingApp()
    window = qtmodern.windows.ModernWindow(window)
    window.show()

    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
