import pandas as pd
import os
import re
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QGridLayout, QLineEdit, QMessageBox, QFileDialog, QTextEdit, QLabel, QDesktopWidget, QListWidget, QListWidgetItem
from PyQt5.QtGui import QFont, QColor
import qtmodern.styles
import qtmodern.windows
from PyQt5 import QtGui
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from PyQt5.QtWidgets import QInputDialog
from openpyxl.utils import get_column_letter


app = QApplication([])
font = QFont("微軟正黑體", 9)
app.setFont(font)


class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Map Defect editor")
        app.setWindowIcon(QtGui.QIcon('hzdjd.ico'))
        self.layout = QGridLayout()

        self.Defect_code_mappings = {}
        self.load_Defect_code_mappings()

        self.undefined_codes = set()

        self.folder_button = QPushButton("點擊輸入貨批資訊")
        self.folder_button.clicked.connect(self.choose_folder)
        self.layout.addWidget(self.folder_button, 0, 0, 1, 2)

        self.folder_line_edit = QLineEdit()
        self.folder_line_edit.setReadOnly(True)
        self.layout.addWidget(self.folder_line_edit, 1, 0, 1, 2)

        self.execute_button = QPushButton("執行")
        self.execute_button.clicked.connect(self.execute_program)
        self.layout.addWidget(self.execute_button, 2, 0, 1, 2)

        self.Defect_code_list = QListWidget()
        self.Defect_code_list.itemClicked.connect(self.list_item_clicked)
        self.layout.addWidget(self.Defect_code_list, 3, 0, 1, 2)

        self.mapping_input_line = QLineEdit()
        # self.layout.addWidget(self.mapping_input_line, 4, 0)

        # self.add_mapping_button = QPushButton("新增")
        # self.add_mapping_button.clicked.connect(self.add_mapping)
        # self.layout.addWidget(self.add_mapping_button, 4, 1)

        # self.remove_mapping_button = QPushButton("刪除")
        # self.remove_mapping_button.clicked.connect(self.remove_mapping)
        # self.layout.addWidget(self.remove_mapping_button, 5, 1)

        self.search_label = QLabel("搜尋Defect code:")
        self.layout.addWidget(self.search_label, 6, 0)

        self.search_line_edit = QLineEdit()
        self.layout.addWidget(self.search_line_edit, 7, 0)

        self.search_button = QPushButton("搜尋")
        self.search_button.clicked.connect(self.search_mapping)
        self.layout.addWidget(self.search_button, 7, 1)

        self.update_Defect_code_list()

        self.setLayout(self.layout)

        self.group = ""
        self.station = ""
        self.batch = ""

    def list_item_clicked(self, item):
        self.mapping_input_line.setText(item.text())

    def load_Defect_code_mappings(self):
        try:
            df = pd.read_excel("Defect_code.xlsx")  # 讀取.xlsx文件
            for _, row in df.iterrows():
                self.Defect_code_mappings[str(row['Key'])] = str(row['Value'])
        except FileNotFoundError:
            QMessageBox.warning(self, "警告", "未找到'Defect_code.xlsx'文件")
            self.Defect_code_mappings = {
                '241': '999',
                '7': '100',
                '9': '102',
                '16': '200'
            }

    def update_Defect_code_list(self):
        self.Defect_code_list.clear()
        for key, value in self.Defect_code_mappings.items():
            item = QListWidgetItem(f"{key},{value}")
            self.Defect_code_list.addItem(item)

    def add_mapping(self):
        mapping_input = self.mapping_input_line.text()
        match = re.match(r"(\d+),(\d+)", mapping_input)
        if match:
            key = match.group(1)
            value = match.group(2)
            self.Defect_code_mappings[key] = value
            self.update_Defect_code_list()
        else:
            QMessageBox.warning(self, "錯誤", "輸入格式不正確，必須是兩個數字用逗號分隔")

    def remove_mapping(self):
        mapping_input = self.mapping_input_line.text()
        match = re.match(r"(\d+),(\d+)", mapping_input)
        if match:
            key = match.group(1)
            value = match.group(2)
            if key in self.Defect_code_mappings and self.Defect_code_mappings[key] == value:
                del self.Defect_code_mappings[key]
                self.update_Defect_code_list()
            else:
                QMessageBox.warning(self, "錯誤", "找不到要刪除的映射")
        else:
            QMessageBox.warning(self, "錯誤", "輸入格式不正確，必須是兩個數字用逗號分隔")

    def search_mapping(self):
        search_input = self.search_line_edit.text()
        if search_input in self.Defect_code_mappings:
            QMessageBox.information(
                self, "搜尋結果", f"此Defect code參數設定為{search_input},{self.Defect_code_mappings[search_input]}")
        else:
            QMessageBox.information(self, "搜尋結果", "此Defect code沒定義參數")

    def choose_folder(self):
        default_path = 'K:\\ALL scanresults'
        try:
            self.group, ok1 = QInputDialog.getText(self, "選擇資料夾", "輸入Group:(ex:BN59、CC60、AB29)")
            self.station, ok2 = QInputDialog.getText(self, "選擇資料夾", "輸入站點:(ex:IQC、AVI、VM)")
            self.batch, ok3 = QInputDialog.getText(self, "選擇資料夾", "輸入客批(Customer Lot ID):")

            if not (ok1 and ok2 and ok3):
                return

            # Rule for station
            self.station = self.station.upper() if self.station.upper() not in ['AVI', 'VM', 'V/M','avi','vm','v/m'] else 'FQC'
            self.station = self.station.upper() if self.station.upper() not in ['IQC', 'AVI_IQC','AVI IQC','iqc','avi_iqc','avi iqc'] else 'IQC'
            self.group = self.group.upper()
            self.batch = self.batch.upper()

            # Find the matching folder
            for folder in os.listdir(default_path):
                folder_parts = folder.split('-', 4)
                if len(folder_parts) >= 4:
                    folder_group = folder_parts[1].strip()
                    folder_station = folder_parts[2].strip()
                    if folder_group == self.group and folder_station == self.station:
                        subfolder_path = os.path.join(default_path, folder)
                        for subfolder in os.listdir(subfolder_path):
                            if subfolder == self.batch:
                                folder_path = os.path.join(subfolder_path, subfolder)
                                self.folder_line_edit.setText(folder_path)
                                return

        except Exception as e:
            print(f"Error occurred: {e}")
            QMessageBox.warning(self, "警告", "出現異常")

        QMessageBox.warning(self, "警告", "未找到符合的資料夾")

    def execute_program(self):
        self.execute_button.setEnabled(False)
        folder_path = self.folder_line_edit.text()

        df = pd.DataFrame(list(self.Defect_code_mappings.items()), columns=['Key', 'Value'])

        # Save the dataframe to an Excel file with custom font and column widths
        self.save_excel_file(df, "Defect_code.xlsx")  

        self.update_Defect_code_list()

        self.update_Defect_code_list()

        if folder_path:
            self.process_folder(folder_path)
            if self.undefined_codes:
                QMessageBox.warning(
                    self, "錯誤", f"code {','.join(self.undefined_codes)} 沒有設定Defect code替換值，請確認")
            QMessageBox.information(self, "完成", "程式已執行完成")
        else:
            QMessageBox.information(self, "更新完成", "Defect code已更新，請重新開啟程式")

    def process_folder(self, folder_path):
        folder_name = os.path.basename(folder_path)
        subfolders = []

        for subfolder in os.listdir(folder_path):
            if os.path.isdir(os.path.join(folder_path, subfolder)):
                if subfolder in ['{:02}'.format(i) for i in range(1, 26)]:
                    subfolders.append(subfolder)

        subfolders.sort()
        self.copy_file_names(subfolders, folder_path)
        self.modify_txt_files(folder_path)

    def copy_file_names(self, subfolders, folder_path):
        txt_folder = os.path.join('K:\\Defect code ink out座標', self.batch)
        os.makedirs(txt_folder, exist_ok=True)

        for subfolder in subfolders:
            subfolder_path = os.path.join(folder_path, subfolder)
            txt_file_path = os.path.join(txt_folder, f"{subfolder}.txt")

            with open(txt_file_path, "w") as txt_file:
                for file_name in os.listdir(subfolder_path):
                    if file_name.lower().endswith((".jpeg", ".jpg", ".png")):
                        txt_file.write(file_name + "\n")

    def modify_txt_files(self, folder_path):
        txt_folder_path = os.path.join('K:\\Defect code ink out座標', self.batch)
        for txt_file_name in os.listdir(txt_folder_path):
            txt_file_path = os.path.join(txt_folder_path, txt_file_name)
            if txt_file_name.lower().endswith(".txt"):
                df = self.modify_txt_file(txt_file_path)
                xlsx_file_path = txt_file_path.replace(".txt", ".xlsx")
                self.save_excel_file(df, xlsx_file_path)

                # Delete the generated .txt file
                os.remove(txt_file_path)

    def modify_txt_file(self, txt_file_path):
        data = []
        with open(txt_file_path, "r") as txt_file:
            for line in txt_file:
                file_name = line.strip()
                match = re.match(r'.*_(\d+)_(\d+)_(\d+)(?:_|\.).*', file_name)
                if match:
                    x_coord = match.group(1)
                    y_coord = match.group(2)
                    defect_code = match.group(3)

                    if defect_code in self.Defect_code_mappings:
                        defect_code = self.Defect_code_mappings[defect_code]
                    else:
                        self.undefined_codes.add(defect_code)

                    map_out_defect_code = f"{x_coord},{y_coord},{defect_code}"
                    data.append([x_coord, y_coord, defect_code, map_out_defect_code])

        df = pd.DataFrame(data, columns=['x_coord', 'y_coord', 'Defect_code', 'Map out defect code'])
        return df

    def save_excel_file(self, df, file_path):
        workbook = Workbook()
        worksheet = workbook.active

        # Write the dataframe to the worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            worksheet.append(r)

        # Set font and size for all cells in the worksheet
        font = Font(name='微軟正黑體', size=9)
        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = font

        # Set column widths
        column_widths = {'A': 19.29, 'B': 19.29, 'C': 19.29, 'D': 19.29}
        for col, width in column_widths.items():
            worksheet.column_dimensions[get_column_letter(ord(col) - 64)].width = width

        # Save the workbook
        workbook.save(file_path)

window = MainWindow()
qtmodern.styles.dark(app)
mw = qtmodern.windows.ModernWindow(window)

screen = QDesktopWidget().screenGeometry()
window_geometry = mw.geometry()
x = int((screen.width() - window_geometry.width()) / 2)
y = int((screen.height() - window_geometry.height()) / 2)
mw.move(x, y)

mw.show()
mw.resize(800, 600)
window.show()

app.exec_()
