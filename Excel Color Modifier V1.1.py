import sys
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QPushButton, QFileDialog, QLabel, QMessageBox
from PyQt5.QtGui import QFont, QIcon
import qtmodern.styles
import qtmodern.windows
import openpyxl
from openpyxl.styles import PatternFill, Border

class ImageClassifier(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle('Excel Color Modifier')
        self.setWindowIcon(QIcon('icon_result.ico'))
        self.resize(400, 200)  
        self.move(600, 800)

        layout = QVBoxLayout()

        self.selectFileBtn = QPushButton('選擇檔案')
        self.selectFileBtn.clicked.connect(self.openFileNameDialog)
        layout.addWidget(self.selectFileBtn)

        self.filePathLabel = QLabel('未選擇檔案')
        layout.addWidget(self.filePathLabel)

        self.runBtn = QPushButton('執行')
        self.runBtn.clicked.connect(self.runModification)
        layout.addWidget(self.runBtn)

        self.setLayout(layout)

    def openFileNameDialog(self):
        options = QFileDialog.Options()
        fileName, _ = QFileDialog.getOpenFileName(self, "選擇 Excel 檔案", "", "Excel Files (*.xlsx)", options=options)
        if fileName:
            self.selectedFile = fileName
            self.filePathLabel.setText(f'選擇的檔案: {fileName}')

    def runModification(self):
        if hasattr(self, 'selectedFile'):
            try:
                self.changeCellColors(self.selectedFile)
                self.modifyExcelFile(self.selectedFile)
                QMessageBox.information(self, "完成", "檔案處理完成！")
            except Exception as e:
                QMessageBox.warning(self, "失敗", f"處理失敗：{e}")

    def rgb_to_hex(self, r, g, b):
        return f"{r:02x}{g:02x}{b:02x}"

    def changeCellColors(self, file_path):
        wb = openpyxl.load_workbook(file_path)
        light_blue_hex = self.rgb_to_hex(198, 226, 255)

        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.fill.start_color.index not in ['00000000', light_blue_hex]:
                        cell.fill = PatternFill(start_color=light_blue_hex, fill_type='solid')

        wb.save(file_path)

    def modifyExcelFile(self, file_path):
        wb = openpyxl.load_workbook(file_path)

        for sheet in wb.sheetnames:
            ws = wb[sheet]

            for row in ws.iter_rows(min_row=12, max_row=50, min_col=1, max_col=2):
                for cell in row:
                    cell.value = None
                    cell.fill = PatternFill(fill_type=None)
                    cell.border = Border()

            color_count = {}
            for row in ws.iter_rows():
                for cell in row:
                    color = cell.fill.start_color.index
                    if color != '00000000':
                        color_count[color] = color_count.get(color, 0) + 1

            row_num = 12
            for color, count in color_count.items():
                ws.cell(row=row_num, column=1).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                ws.cell(row=row_num, column=2).value = count
                row_num += 1

        wb.save(file_path)

if __name__ == '__main__':
    app = QApplication(sys.argv)
    font = QFont("微軟正黑體", 9)
    app.setFont(font)
    qtmodern.styles.dark(app)

    window = ImageClassifier()
    win = qtmodern.windows.ModernWindow(window)
    win.show()

    sys.exit(app.exec_())
