import os
import shutil
import openpyxl
from openpyxl.styles import Font, Color, PatternFill, Alignment

# 取得桌面路徑
desktop_path = os.path.expanduser("~/Desktop")
# 指定檔案名稱
file_name = "standard_deviation_for_confidence.xlsx"
# 組合完整的檔案路徑
file_path = os.path.join(desktop_path, file_name)
# 開啟Excel檔案
workbook = openpyxl.load_workbook(file_path)
# 選取要設定欄寬的工作表（Sheet）
workSheet = workbook['Sheet']
# 將A欄的欄寬設定為42
workSheet.column_dimensions['A'].width = 53
# 將B欄到Z欄的欄寬設定為20
for column_letter in range(ord('B'), ord('Z')+1):
    column_letter = chr(column_letter)
    workSheet.column_dimensions[column_letter].width = 20
# 新增字體設定為微軟正黑體
font = Font(name='微軟正黑體')
for row in workSheet.iter_rows():
    for cell in row:
        cell.font = font
# 新增將D欄到Z欄的所有儲存格轉換成數字
for row in workSheet.iter_rows(min_row=2, max_row=workSheet.max_row, min_col=4, max_col=26):
    for cell in row:
        try:
            cell.value = float(cell.value)
        except (ValueError, TypeError):
            # 如果無法轉換為數字，保留原始值
            pass
# 新增將D欄到Z欄的所有儲存格文字置中
for row in workSheet.iter_rows(min_row=2, max_row=workSheet.max_row, min_col=4, max_col=26):
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')
# 儲存變更
workbook.save(file_path)
# 關閉Excel檔案
workbook.close()





# 取得桌面路徑
desktop_path = os.path.expanduser("~/Desktop")

# 複製檔案到新的檔名
shutil.copy(os.path.join(desktop_path, "standard_deviation_for_confidence.xlsx"), os.path.join(desktop_path, "pass.xlsx"))
shutil.copy(os.path.join(desktop_path, "standard_deviation_for_confidence.xlsx"), os.path.join(desktop_path, "ntc.xlsx"))

# 要建立的資料夾路徑
desktop_path = os.path.expanduser("~/Desktop")
folder_name = "standard_deviation_for_confidence"
folder_path = os.path.join(desktop_path, folder_name)

# 如果資料夾不存在，則建立它
if not os.path.exists(folder_path):
    os.makedirs(folder_path)

# 要移動的檔案清單
files_to_move = ["standard_deviation_for_confidence.xlsx", "standard_deviation_for_confidence.txt", "pass.xlsx", "ntc.xlsx"]

# 移動檔案到新資料夾
for file_name in files_to_move:
    source_path = os.path.join(desktop_path, file_name)
    destination_path = os.path.join(folder_path, file_name)
    
    # 檢查檔案是否存在，然後執行剪切操作
    if os.path.exists(source_path):
        shutil.move(source_path, destination_path)
        print(f"已剪切檔案 {file_name} 到 {folder_name} 資料夾。")
    else:
        print(f"檔案 {file_name} 不存在於桌面。")

print("完成所有操作。")







# 1. 讀取Excel檔案(ntc，輸出最大值)
# 檔案路徑
file_path = os.path.expanduser("~/Desktop/standard_deviation_for_confidence/ntc.xlsx")

# 讀取Excel檔案
wb = openpyxl.load_workbook(file_path)
Sheet = wb['Sheet']

# 獲取工作表的最大行數
max_row = Sheet.max_row

# 從最後一行往前遍歷，以防止刪除行時導致索引錯誤
for row_num in range(max_row, 1, -1):
    # 獲取B列和C列的值
    b_value = Sheet.cell(row=row_num, column=2).value
    c_value = Sheet.cell(row=row_num, column=3).value
    
    # 如果B列的值等於C列的值，則刪除該列
    if b_value == c_value:
        Sheet.delete_rows(row_num)


# 2. 設定RGB顏色
rgb_color = "E7B1B2"  # 231, 177, 178的RGB顏色的十六進制表示
fill = PatternFill(start_color=rgb_color, end_color=rgb_color, fill_type="solid")

# 3. 將D欄~Z欄中每欄中的最大值設為指定的RGB顏色，同時檢查每列是否有該顏色，並設定A欄的顏色
for column in range(4, 27):  # D欄~Z欄的欄位編號
    max_value = float('-inf')  # 初始化最大值為負無窮大
    for row in range(2, Sheet.max_row + 1):
        cell_value = Sheet.cell(row=row, column=column).value
        if cell_value is not None and cell_value > max_value:
            max_value = cell_value
    # 設定欄位的填充顏色
    for row in range(2, Sheet.max_row + 1):
        if Sheet.cell(row=row, column=column).value == max_value:
            Sheet.cell(row=row, column=column).fill = fill
            Sheet.cell(row=row, column=1).fill = fill  # 設定A欄的填充顏色
#for column in range(4, 27):  # D欄~Z欄的欄位編號
#    min_value = float('inf')  # 初始化最小值為正無窮大
#    for row in range(2, Sheet.max_row + 1):
#        cell_value = Sheet.cell(row=row, column=column).value
#        if cell_value is not None and cell_value < min_value:
#            min_value = cell_value
#    # 設定欄位的填充顏色
#    for row in range(2, Sheet.max_row + 1):
#        if Sheet.cell(row=row, column=column).value == min_value:
#            Sheet.cell(row=row, column=column).fill = fill
#            Sheet.cell(row=row, column=1).fill = fill  # 設定A欄的填充顏色

# 4. 新增，啟用篩選功能，並設定A欄中有顏色的列為篩選條件
Sheet.auto_filter.ref = Sheet.dimensions

# 5. 隱藏A欄中沒有顯示儲存格顏色的列
for row in range(2, Sheet.max_row + 1):
    if Sheet.cell(row=row, column=1).fill != fill:
        Sheet.row_dimensions[row].hidden = True

# 6. 儲存修改後的檔案
wb.save(file_path)
wb.close()

print("已完成設定RGB顏色！")








# 1. 讀取Excel檔案(pass，輸出最大值)
files_path = os.path.expanduser("~/Desktop/standard_deviation_for_confidence/pass.xlsx")
wb = openpyxl.load_workbook(files_path)
Sheet = wb['Sheet']

# 獲取工作表的最大行數
max_row = Sheet.max_row

# 從最後一行往前遍歷，以防止刪除行時導致索引錯誤
for row_num in range(max_row, 1, -1):
    # 獲取B列和C列的值
    b_value = Sheet.cell(row=row_num, column=2).value
    c_value = Sheet.cell(row=row_num, column=3).value
    
    # 如果B列的值等於C列的值，則刪除該列
    if b_value == c_value:
        Sheet.delete_rows(row_num)

# 2. 設定RGB顏色
rgb_color = "E7B1B2"  # 231, 177, 178的RGB顏色的十六進制表示
fill = PatternFill(start_color=rgb_color, end_color=rgb_color, fill_type="solid")

# 3. 將D欄~Z欄中每欄中的最大值設為指定的RGB顏色，同時檢查每列是否有該顏色，並設定A欄的顏色
for column in range(4, 27):  # D欄~Z欄的欄位編號
    max_value = 0  # 初始化最大值為0
    for row in range(2, Sheet.max_row + 1):
        cell_value = Sheet.cell(row=row, column=column).value
        if cell_value is not None and cell_value > max_value:
            max_value = cell_value
    # 設定欄位的填充顏色
    for row in range(2, Sheet.max_row + 1):
        if Sheet.cell(row=row, column=column).value == max_value:
            Sheet.cell(row=row, column=column).fill = fill
            Sheet.cell(row=row, column=1).fill = fill  # 設定A欄的填充顏色

# 4. 新增，啟用篩選功能，並設定A欄中有顏色的列為篩選條件
Sheet.auto_filter.ref = Sheet.dimensions

# 5. 隱藏A欄中沒有顯示儲存格顏色的列
for row in range(2, Sheet.max_row + 1):
    if Sheet.cell(row=row, column=1).fill != fill:
        Sheet.row_dimensions[row].hidden = True

# 6. 儲存修改後的檔案
wb.save(files_path)
wb.close()

print("已完成設定RGB顏色！")

# 1. 讀取桌面proc_option.txt裡面的值
desktop_path = os.path.expanduser("~/Desktop")
file_path = os.path.join(desktop_path, "proc_option.txt")

try:
    with open(file_path, 'r') as file:
        option = file.read().strip()
except FileNotFoundError:
    print("找不到proc_option.txt檔案")
    option = ""

# 2. 當裡面的值是proc.ntc，就將桌面內名稱為standard_deviation_for_confidence資料夾裡面的pass.xlsx檔案刪除
if option == "proc.ntc":
    folder_path = os.path.join(desktop_path, "standard_deviation_for_confidence")
    pass_excel_path = os.path.join(folder_path, "pass.xlsx")
    
    try:
        os.remove(pass_excel_path)
        print("已刪除 pass.xlsx 檔案")
    except FileNotFoundError:
        print("找不到 pass.xlsx 檔案")

# 3. 當裡面的值是pass.ntc，就將桌面內名稱為standard_deviation_for_confidence資料夾裡面的ntc.xlsx檔案刪除
elif option == "proc.pass":
    folder_path = os.path.join(desktop_path, "standard_deviation_for_confidence")
    ntc_excel_path = os.path.join(folder_path, "ntc.xlsx")
    
    try:
        os.remove(ntc_excel_path)
        print("已刪除 ntc.xlsx 檔案")
    except FileNotFoundError:
        print("找不到 ntc.xlsx 檔案")

# 4. 完成後刪除桌面的proc_option.txt
try:
    os.remove(file_path)
    print("已刪除 proc_option.txt 檔案")
except FileNotFoundError:
    print("找不到 proc_option.txt 檔案")

# 開啟桌面內名稱為standard_deviation_for_confidence資料夾供user看
os.startfile(folder_path)